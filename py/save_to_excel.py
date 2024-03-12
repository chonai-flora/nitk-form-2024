import argparse
import openpyxl
import datetime
import zoneinfo
import collections
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side

from firestore_client import FirestoreClient


course_ids = [
    "A",
    "B",
    "C",
    "D",
    "E",
    "F",
]
titles = [
    "DNAストラップ-星形ストラップ",
    "メタル飛行機",
    "ピンホール・レンズカメラ",
    "ふしぎな水そう",
    "球体ロボットでプログラミング体験",
    "電卓を分解して自分だけのアクセサリーを作ろう",
]
key_order = [
    "title",
    "schedule",
    "name",
    "grade",
    "email",
    "submitted_at",
]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--schedule", type=str)
    args = parser.parse_args()

    # エントリーポイント
    date = datetime.datetime.now(zoneinfo.ZoneInfo("Asia/Tokyo"))
    headers = [
        f"{date:%Y/%m/%d %H:%M更新(自動作成)}",
        "時間",
        "名前",
        "学年",
        "メールアドレス",
        "受付完了時刻",
    ]

    # collectionの取得
    firestore_client = FirestoreClient()
    ref = firestore_client.fetch_ref("applicants")

    # フォント, 罫線, カラー
    font = Font(name="游ゴシック")
    thin = Side(style="thin", color="000000")
    dashed = Side(style="dashed", color="000000")
    fill = PatternFill(patternType="solid", fgColor="ebf1de")

    # Excelファイルの初期化
    wb = openpyxl.Workbook()
    wb.remove(wb["Sheet"])

    # データの書き込み
    for index, (course_id, title) in enumerate(zip(course_ids, titles)):
        # シートの追加
        wb.create_sheet(index=index, title=title)
        ws = wb[title]

        # ヘッダーの書き込み
        for col, header in enumerate(headers):
            cell = ws.cell(row=1, column=col+1, value=header)
            cell.font = font
            cell.fill = fill
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

        # DBにアクセス
        docs = ref.document(course_id).get().to_dict()
        ordered_docs = collections.OrderedDict(sorted(docs.items()))
        if args.schedule is not None and args.schedule in ordered_docs:
            items = dict([(args.schedule, ordered_docs.get(args.schedule))]).items()
        else:
            items = ordered_docs.items()

        row = 2
        title_written = False        
        for schedule, users in items:
            if not users:
                continue

            schedule_written = False

            for user in users:
                if title_written:
                    title = ""
                else:
                    title_written = True

                if schedule_written:
                    schedule = ""
                else:
                    schedule_written = True
                
                user["title"], user["schedule"] = title.replace('-', '/'), schedule
                for col, key in enumerate(key_order):
                    border = Border(top=thin, bottom=thin, left=dashed, right=dashed)

                    if key == "title" or key == "schedule":
                        border = Border(left=thin, right=thin)
                    elif key == "name":
                        border = Border(top=thin, bottom=thin, left=thin, right=dashed)
                    elif key == "submitted_at":
                        submitted_at = datetime.datetime.fromtimestamp(user[key].timestamp(), tz=zoneinfo.ZoneInfo("Asia/Tokyo"))
                        user[key] = f"{submitted_at:%Y-%m-%d %H:%M:%S}"
                        border = Border(top=thin, bottom=thin, left=dashed, right=thin)

                    cell = ws.cell(row=row, column=col+1, value=user[key])
                    cell.font = font
                    cell.border = border
                
                row += 1
            
            ws.cell(row=row-1, column=2).border = Border(bottom=thin, left=thin, right=dashed)

        ws.cell(row=row-1, column=1).border = Border(bottom=thin, left=thin, right=thin)

    # 保存
    if args.schedule is None:
        wb.save("./applicants/申込者リスト.xlsx")
    else:
        wb.save(f"./applicants/{args.schedule.replace(':', '')}申込者リスト.xlsx")


if __name__ == "__main__":
    main()